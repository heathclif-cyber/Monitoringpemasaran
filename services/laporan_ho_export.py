"""Export laporan ke format HO — hanya mengisi bagian Penjualan Lokal (baris 71–99)."""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from services.ba_utils import ba_effective_harga, calculate_ba_pokok, is_payung_ba

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "laporan_ho_template.xlsx"

MONTHS_ID = [
    "",
    "Januari",
    "Februari",
    "Maret",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Agustus",
    "September",
    "Oktober",
    "November",
    "Desember",
]

# Baris komoditas di sheet template (kolom C = uraian)
LOKAL_ROW_BY_KEY: dict[str, int] = {
    "minyak_sawit": 71,
    "inti_sawit": 72,
    "pko": 73,
    "pkm": 74,
    "tbs": 75,
    "karet_rss1": 77,
    "karet_sir10": 78,
    "karet_sir20": 79,
    "karet_lateks": 80,
    "karet_lainnya": 81,
    "teh_ctc": 83,
    "teh_orthodox": 84,
    "teh_hijau": 85,
    "teh_lainnya": 86,
    "gula": 87,
    "tetes": 88,
    "kopi_arabika": 90,
    "kopi_robusta": 91,
    "kakao_edel": 93,
    "kakao_bulk": 94,
    "kayu": 95,
    "tembakau": 96,
    "hortikultura": 97,
    "forestry": 98,
    "lain_lain": 99,
}

LOKAL_DATA_ROWS = sorted(LOKAL_ROW_BY_KEY.values())


@dataclass
class HoBucket:
    kuantum: float = 0.0
    nilai: float = 0.0

    def add(self, kuantum: float, nilai: float) -> None:
        self.kuantum += kuantum
        self.nilai += nilai

    @property
    def per_unit(self) -> float:
        return self.nilai / self.kuantum if self.kuantum > 0 else 0.0


def _norm(text: str | None) -> str:
    return (text or "").strip().lower()


def map_komoditi_to_ho_key(komoditi: str | None, deskripsi: str | None) -> str:
    k = _norm(komoditi)
    d = _norm(deskripsi)
    combined = f"{k} {d}"

    if any(x in combined for x in ("minyak sawit", "cpo", "crude palm", "sawit mentah")):
        return "minyak_sawit"
    if any(x in combined for x in ("inti sawit", "kernel", "intan sawit", "pk")) and "pko" not in combined and "pkm" not in combined:
        return "inti_sawit"
    if "pko" in combined or "palm kernel oil" in combined:
        return "pko"
    if any(x in combined for x in ("pkm", "pke", "palm kernel meal", "bungkil")):
        return "pkm"
    if any(x in combined for x in ("tbs", "tandan buah segar", "fresh fruit bunch")):
        return "tbs"

    if "karet" in k or "rubber" in k or "rss" in combined or "sir" in combined or "latek" in combined or "latex" in combined or "crepe" in combined or "lump" in combined:
        if "rss" in combined and ("1" in combined or "rss1" in combined or "rss 1" in combined):
            return "karet_rss1"
        if "sir 10" in combined or "sir10" in combined:
            return "karet_sir10"
        if "sir 20" in combined or "sir20" in combined:
            return "karet_sir20"
        if "latek" in combined or "latex" in combined:
            return "karet_lateks"
        return "karet_lainnya"

    if "teh" in k or "tea" in k:
        if "ctc" in combined:
            return "teh_ctc"
        if "orthodox" in combined:
            return "teh_orthodox"
        if "hijau" in combined or "green" in combined:
            return "teh_hijau"
        return "teh_lainnya"

    if any(x in combined for x in ("tetes", "molasses", "vit c")):
        return "tetes"

    # Tebu (komoditi) + material gula (gapoktan, kemasan PG, dll.) → baris Gula
    if (
        "gula" in k
        or "sugar" in k
        or "tebu" in k
        or "gula" in d
        or "gapoktan" in combined
        or "kemasan" in d and "kg" in d
    ):
        return "gula"

    if "kopi" in k or "coffee" in k:
        if "arabika" in combined or "arabica" in combined:
            return "kopi_arabika"
        if "robusta" in combined:
            return "kopi_robusta"
        return "kopi_robusta"

    if "kakao" in k or "cocoa" in k or "cacao" in k:
        if "edel" in combined:
            return "kakao_edel"
        if "bulk" in combined:
            return "kakao_bulk"
        return "kakao_bulk"

    if "kayu" in k or "wood" in k or "timber" in k:
        return "kayu"
    if "tembakau" in k or "tobacco" in k:
        return "tembakau"
    if any(x in combined for x in ("hortikultura", "horti", "sayur", "buah")):
        return "hortikultura"
    if "forestry" in combined or "hutan" in k:
        return "forestry"

    if "kelapa" in k or "kopra" in k:
        return "lain_lain"
    if "sawit" in k:
        return "tbs"

    return "lain_lain"


def _format_bulan_buku(date_val) -> str:
    if not date_val:
        return ""
    try:
        name = MONTHS_ID[date_val.month]
    except (AttributeError, IndexError, TypeError):
        return ""
    return f"{date_val.month:02d}-{name}"


def _find_invoice_for_ba(db, ba) -> object | None:
    import models

    linked = getattr(ba, "invoice", None)
    if linked:
        return linked
    no_ba = ba.no_ba
    return (
        db.query(models.Invoice)
        .filter(
            (models.Invoice.no_ba == no_ba) | (models.Invoice.no_kontrak == no_ba)
        )
        .first()
    )


def _ba_pendapatan_pokok(ba, kontrak, invoice=None) -> int:
    """Nilai pokok BA — fallback ke invoice jika harga BA/kontrak payung belum terisi."""
    volume = float(ba.volume_ba or 0)
    harga = ba_effective_harga(ba, kontrak)
    pokok = round(calculate_ba_pokok(kontrak, volume, harga))
    if pokok > 0:
        return pokok

    inv = invoice or getattr(ba, "invoice", None)
    if inv:
        inv_gross = float(inv.jumlah_pembayaran or 0)
        if inv_gross > 0:
            if str(getattr(kontrak, "is_ppn", "true")).lower() != "false":
                ppn_rate = float(getattr(kontrak, "ppn_persen", 0) or 0) / 100
                if ppn_rate > 0:
                    return round(inv_gross / (1 + ppn_rate))
            return round(inv_gross)

    if volume > 0 and harga > 0:
        return round(volume * harga)
    return 0


def _ba_to_ho_row(ba, kontrak, invoice=None) -> dict:
    """Baris penjualan HO dari Berita Acara — diakui terjual saat barang dikirim."""
    volume = float(ba.volume_ba or 0)
    pokok = _ba_pendapatan_pokok(ba, kontrak, invoice=invoice)
    buku = ba.bulan_buku or ba.tanggal_ba
    komoditi = ba.komoditi or kontrak.komoditi or ""
    deskripsi = ba.deskripsi or kontrak.jenis_komoditi or kontrak.deskripsi_produk or komoditi
    unit = ba.nama_unit or kontrak.kebun_produsen or ""

    return {
        "Komoditi": komoditi,
        "Deskripsi_Produk": deskripsi,
        "Unit": unit,
        "Jumlah_DO": volume,
        "DPP_Pokok": pokok,
        "Pendapatan_Pokok": pokok,
        "Satuan": kontrak.satuan or "Kg",
        "No_BA": ba.no_ba,
        "Tanggal_BA": ba.tanggal_ba.strftime("%Y-%m-%d") if ba.tanggal_ba else "",
        "Bulan_Buku": _format_bulan_buku(buku),
        "Rencana_Pengambilan": ba.tanggal_ba.strftime("%Y-%m-%d") if ba.tanggal_ba else "",
        "Raw_Date": buku.strftime("%Y-%m-%d") if buku else "",
        "Is_Payung_BA": True,
    }


def fetch_payung_ba_ho_rows(
    db,
    *,
    units: list[str] | None = None,
    komoditis: list[str] | None = None,
) -> list[dict]:
    """Ambil semua BA kontrak payung untuk pengakuan penjualan HO."""
    from sqlalchemy.orm import joinedload

    import models

    unit_set = set(units or [])
    komoditi_set = set(komoditis or [])

    bas = (
        db.query(models.BeritaAcara)
        .options(
            joinedload(models.BeritaAcara.kontrak),
            joinedload(models.BeritaAcara.invoice),
        )
        .all()
    )

    rows: list[dict] = []
    for ba in bas:
        kontrak = ba.kontrak
        if not kontrak or not is_payung_ba(kontrak):
            continue
        if float(ba.volume_ba or 0) <= 0:
            continue

        invoice = _find_invoice_for_ba(db, ba)
        ho_row = _ba_to_ho_row(ba, kontrak, invoice=invoice)
        if unit_set and ho_row.get("Unit") not in unit_set:
            continue
        if komoditi_set and ho_row.get("Komoditi") not in komoditi_set:
            continue
        rows.append(ho_row)

    return rows


def _merge_payung_ba_row(laporan_row: dict, ba_row: dict) -> dict:
    """Volume dari BA (pengiriman); nilai mengikuti Laporan Digital (Pendapatan_Pokok)."""
    merged = dict(ba_row)
    pokok = float(laporan_row.get("Pendapatan_Pokok") or 0)
    if pokok > 0:
        merged["Pendapatan_Pokok"] = pokok
        merged["DPP_Pokok"] = float(laporan_row.get("DPP_Pokok") or pokok)
    if float(laporan_row.get("Jumlah_DO") or 0) <= 0:
        merged["Jumlah_DO"] = float(ba_row.get("Jumlah_DO") or 0)
    return merged


def prepare_ho_export_rows(client_rows: list[dict], ba_supplement: list[dict]) -> list[dict]:
    """
    Kontrak payung: penjualan HO mengikuti BA (pengiriman), bukan menunggu DO/transfer.
    - Jika sudah ada DO dengan volume → pakai baris DO.
    - Jika belum ada DO → volume dari BA, nilai tetap Pendapatan_Pokok laporan.
    """
    ba_by_no = {r["No_BA"]: r for r in ba_supplement if r.get("No_BA")}

    do_volume_by_ba: dict[str, float] = {}
    rows_by_ba: dict[str, list[dict]] = {}
    non_ba_rows: list[dict] = []

    for row in client_rows:
        no_ba = (row.get("No_BA") or "").strip()
        if not no_ba:
            non_ba_rows.append(row)
            continue
        rows_by_ba.setdefault(no_ba, []).append(row)
        do_volume_by_ba[no_ba] = do_volume_by_ba.get(no_ba, 0) + float(row.get("Jumlah_DO") or 0)

    result = list(non_ba_rows)

    for no_ba, grouped in rows_by_ba.items():
        if do_volume_by_ba.get(no_ba, 0) > 0:
            for row in grouped:
                if float(row.get("Jumlah_DO") or 0) > 0:
                    result.append(row)
            continue
        if no_ba in ba_by_no:
            result.append(_merge_payung_ba_row(grouped[0], ba_by_no[no_ba]))
            continue
        for row in grouped:
            result.append(row)
            break

    for no_ba, ba_row in ba_by_no.items():
        if no_ba not in rows_by_ba:
            result.append(ba_row)

    return result


def _extract_year_month(row: dict, mode: str) -> tuple[str, str]:
    # Payung BA: periode HO selalu bulan buku BA (pengiriman), bukan tgl transfer
    if row.get("No_BA") or row.get("Is_Payung_BA"):
        bulan_buku = row.get("Bulan_Buku") or ""
        month = bulan_buku[:2] if len(bulan_buku) >= 2 else ""
        for field in ("Rencana_Pengambilan", "Tanggal_BA", "Raw_Date"):
            raw = row.get(field) or ""
            if len(raw) >= 7 and raw[4] == "-":
                return raw[:4], month
        return "", month

    if mode == "RENCANA":
        rencana = row.get("Rencana_Pengambilan") or ""
        if len(rencana) >= 7:
            return rencana[:4], rencana[5:7]
        bulan_buku = row.get("Bulan_Buku") or ""
        month = bulan_buku[:2] if len(bulan_buku) >= 2 else ""
        raw = row.get("Raw_Date") or row.get("Billing_Date") or ""
        year = raw[:4] if len(raw) >= 4 else ""
        return year, month

    raw = row.get("Raw_Date") or ""
    if len(raw) >= 7:
        return raw[:4], raw[5:7]

    transfer = row.get("Tanggal_Transfer") or ""
    if "/" in transfer:
        parts = transfer.split("/")
        if len(parts) == 3:
            return parts[2], parts[1].zfill(2)

    bulan_buku = row.get("Bulan_Buku") or ""
    month = bulan_buku[:2] if len(bulan_buku) >= 2 else ""
    for field in ("Raw_Date", "Billing_Date"):
        raw = row.get(field) or ""
        if len(raw) >= 4:
            return raw[:4], month
    return "", month


def _row_kuantum_nilai(row: dict) -> tuple[float, float]:
    satuan = _norm(row.get("Satuan") or "kg")
    volume = float(row.get("Jumlah_DO") or 0)
    # Selaras Total Pendapatan Laporan Digital (bukan DPP_Pokok)
    nilai = float(row.get("Pendapatan_Pokok") or row.get("DPP_Pokok") or 0)
    if satuan == "butir":
        return 0.0, nilai
    return volume, nilai


def _aggregate_lokal(
    rows: Iterable[dict],
    *,
    year: str,
    month: str,
    mode: str,
) -> tuple[dict[str, HoBucket], dict[str, HoBucket]]:
    bulan = {key: HoBucket() for key in LOKAL_ROW_BY_KEY}
    ytd = {key: HoBucket() for key in LOKAL_ROW_BY_KEY}
    month_int = int(month) if month.isdigit() else 0

    for row in rows:
        row_year, row_month = _extract_year_month(row, mode)
        if not row_year or row_year != year or not row_month:
            continue
        try:
            row_month_int = int(row_month)
        except ValueError:
            continue

        key = map_komoditi_to_ho_key(row.get("Komoditi"), row.get("Deskripsi_Produk"))
        kuantum, nilai = _row_kuantum_nilai(row)

        if row_month == month:
            bulan[key].add(kuantum, nilai)
        if row_month_int <= month_int:
            ytd[key].add(kuantum, nilai)

    return bulan, ytd


def _month_label(month: str, year: str) -> str:
    try:
        name = MONTHS_ID[int(month)]
    except (ValueError, IndexError):
        name = month
    return f"   Bulan {name} {year}"


def _ytd_label(month: str, year: str) -> str:
    try:
        name = MONTHS_ID[int(month)]
    except (ValueError, IndexError):
        name = month
    return f"   S/d. Bulan {name} {year}"


def _write_bucket(ws, row_idx: int, bucket: HoBucket, *, col_kuantum: int, col_nilai: int) -> None:
    ws.cell(row=row_idx, column=col_kuantum, value=round(bucket.kuantum, 2) if bucket.kuantum else 0)
    ws.cell(row=row_idx, column=col_nilai, value=round(bucket.nilai, 0) if bucket.nilai else 0)


def generate_laporan_ho_xlsx(
    rows: list[dict],
    *,
    year: str,
    month: str,
    mode: str = "TRANSFER",
    db=None,
    filters: dict | None = None,
) -> bytes:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template HO tidak ditemukan: {TEMPLATE_PATH}")

    export_rows = list(rows)
    if db is not None:
        filter_opts = filters or {}
        ba_rows = fetch_payung_ba_ho_rows(
            db,
            units=filter_opts.get("units") or None,
            komoditis=filter_opts.get("komoditis") or None,
        )
        export_rows = prepare_ho_export_rows(export_rows, ba_rows)

    bulan_data, ytd_data = _aggregate_lokal(export_rows, year=year, month=month, mode=mode)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    ws["D3"] = _month_label(month, year)
    ws["G3"] = _ytd_label(month, year)

    for row_idx in LOKAL_DATA_ROWS:
        for col in (4, 6, 7, 9):
            ws.cell(row=row_idx, column=col, value=0)
        for col in (5, 8):
            ws.cell(row=row_idx, column=col, value=None)

    for key, row_idx in LOKAL_ROW_BY_KEY.items():
        _write_bucket(ws, row_idx, bulan_data[key], col_kuantum=4, col_nilai=6)
        _write_bucket(ws, row_idx, ytd_data[key], col_kuantum=7, col_nilai=9)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()